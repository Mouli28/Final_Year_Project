from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from bs4 import BeautifulSoup
import re
import time


def setup_driver():
    """Setup Selenium Chrome driver with default options."""
    options = Options()
    options.add_argument('--disable-gpu')
    options.add_argument('--no-sandbox')
    options.add_argument('--incognito')
    return webdriver.Chrome(options=options)


def get_google_search_urls(query, num_results):
    """
    Retrieves URLs from Google search results for the given query.
    Returns a list of URLs.
    """
    driver = setup_driver()
    search_urls = []
    sites = ['autozone.com','kalpartz.com','shopadvanceautoparts.com','vanhorntruckparts.com','ebay.com','nickstruckparts.com','thewrenchmonkey.com','finditparts.com','www.amazon.com','accessorymods.com','centralalbertapaintsupply','fleetpride.com']
    required_results = (len(sites) * 2) + num_results
    try:
        # Step 1: Search each trusted site individually, taking only 2 URLs per site
        for site in sites:
            site_query = f"{query} {site}"
            driver.get("https://www.google.com")
            search_box = driver.find_element(By.NAME, "q")
            search_box.send_keys(site_query)
            search_box.send_keys(Keys.RETURN)

            # Wait a shorter time for results (5 seconds) to avoid hanging on zero-result pages
            try:
                WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, "div.yuRUbf a")))
                results = driver.find_elements(By.CSS_SELECTOR, "div.yuRUbf a")[:2]
                site_urls = [result.get_attribute("href") for result in results]
                search_urls.extend(site_urls)
            except Exception as e:
                print(f"No results for {site}, moving to the next site.")

            # Stop if we reach the required results
            if len(search_urls) >= required_results:
                break


        # Step 2: Perform a general search if not enough results from trusted sites
        if len(search_urls) < required_results:
            general_query = f"{query}"
            driver.get("https://www.google.com")
            search_box = driver.find_element(By.NAME, "q")
            search_box.send_keys(general_query)
            search_box.send_keys(Keys.RETURN)
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, "div.yuRUbf a")))

            # Extract URLs from the first page of the general search
            results = driver.find_elements(By.CSS_SELECTOR, "div.yuRUbf a")
            general_urls = [result.get_attribute("href") for result in results]
            search_urls.extend(general_urls)

            # Fetch additional pages if needed
            while len(search_urls) < required_results:
                try:
                    next_button = driver.find_element(By.ID, "pnnext")
                    next_button.click()
                    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, "div.yuRUbf a")))
                    results = driver.find_elements(By.CSS_SELECTOR, "div.yuRUbf a")
                    search_urls.extend([result.get_attribute("href") for result in results])
                except Exception as e:
                    print(f"Error navigating to the next page: Endind of the list")
                    break

    finally:
        driver.quit()

    return search_urls[:required_results]
def price_comparison(query, num_results, part_num, description, mfr):
    """
    Performs price comparison for the given query by extracting product details from multiple websites.
    Returns a pandas DataFrame containing the product details.
    """
    urls = get_google_search_urls(query, num_results)
    print(urls)
    #urls = ['https://www.continentalbattery.com/product-reference/advance-diehard-silver-561','https://www.amazon.com/Dorman-485-701-Grease-Fitting/dp/B0036C9DR0','https://excofilter.com/description/AX230207/93800-CARQUEST']
    all_product_details = []
    #urls = [' https://nickstruckparts.com/products/coolantelbow-561-17275-polyester-reinforced-45-deg?srsltid=AfmBOoqwPg0RDxkqbjqBBDDZBwMLwX5sUIjiF4UZarahHiC0qeQUOsXk']
    for url in urls:
        try:
            #print(class_names)
            # Gather product details
            details = {
            'MFR Line' : mfr,
            'Part Num': part_num,
            'Description' : description,
            'url': url,
            'Part No From Site': extract_product_details(url,['part_number', 'partNumSection', 'part-number', 'product_part-info', 'item-part-number','sku','sku number','item part number','part number','item no','product-basic-details__text--part']),
            'Cross Reference': extract_product_details(url,['product-specifications','interchange category-description','cross-reference','other-cross-refrence','product-description rte','productDescription','crossRefWrapper','interchangeItemValue','productDetails_techSpec_section_1','product-details-d','css-cm8roc','product_description__row','product-part-interchange','product-specifications-container','cross-reference-list','product-single__description','x-item-description-child','interchange','Crossreference','replaces', 'product superseded','superseded','interchages']),
            }
            #print(details)
            all_product_details.append(details)
            print(f"Website done: {url}")

        except Exception as e:
            print(f"Error processing URL {url}: {e}")

    return pd.DataFrame(all_product_details)
def extract_product_details(url, substrings):
    # Setup WebDriver
    driver = setup_driver()

    def extract_data_from_soup(soup, substrings):
        """Helper function to extract data based on class names or IDs from a BeautifulSoup object."""
        extracted_data = []
        all_elements = soup.find_all()

        class_names = []
        for element in all_elements:
            classes = element.get('class', [])
            ids = element.get('id', '')
            for cls in classes:
                class_names.append(cls)
            class_names.append(ids)
        #print(class_names)
        for substr in substrings:
            matched_classes = []
            # Normalize class names and IDs by removing spaces, dashes, etc.
            for cls in class_names:
                cls1 = re.sub(r'[\s+\-\_+]', '', cls).lower()
                substr1 = re.sub(r'[\s+\-\_+]', '', substr).lower()
                if substr1 in cls1:
                    matched_classes.append(cls)
            #print(matched_classes)
            # Extract elements based on matching class names or IDs
            for element in all_elements:
                if element.get('id', '') in matched_classes:
                    extracted_text = re.sub(r'\s+', ' ', element.get_text()).strip()
                    extracted_data.append(extracted_text)

                if any(cls in matched_classes for cls in element.get('class', [])):
                    extracted_text = re.sub(r'\s+', ' ', element.get_text()).strip()
                    matched_class_name = ' '.join(element.get('class', []))
                    extracted_data.append(extracted_text)

        return extracted_data

    # Navigate to the URL
    driver.get(url)
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.TAG_NAME, 'body')))

    # Extract data from the main page
    main_soup = BeautifulSoup(driver.page_source, 'html.parser')
    extracted_data = extract_data_from_soup(main_soup, substrings)
    
    # Check for iframes before attempting to extract data from them
    iframes = driver.find_elements(By.TAG_NAME, 'iframe')
    if iframes:
        #print(f"Found {len(iframes)} iframes. Extracting data from them...")

        for iframe in iframes:
            driver.switch_to.frame(iframe)  # Switch to the iframe
            iframe_soup = BeautifulSoup(driver.page_source, 'html.parser')
            iframe_extracted_data = extract_data_from_soup(iframe_soup, substrings)
            extracted_data.extend(iframe_extracted_data)
            driver.switch_to.default_content()  # Switch back to the main document

    # Close the driver
    driver.quit()

    # Print and return extracted details
    #for class_or_id, data in extracted_data:
        #print(f"Class/ID: {class_or_id} | Extracted Data: {data}")

    return extracted_data

if __name__ == "__main__":
    location = r"sites.xlsx"
    list = pd.read_excel(location)
    combined_list = list.apply(lambda row: f"{row['Part num']} {row['Description']} Price", axis=1).tolist()
    for i in range(42,51):
        desc = list.iloc[i]['Description'] 
        partnum = list.iloc[i]['Part num']
        mfr = list.iloc[i]['MFR Line']
        query = combined_list[i]
        df = price_comparison(query, num_results=10, part_num = partnum , description = desc, mfr = mfr)
        file_name = f"{query.replace(' ', '_')}_substr_part_details.xlsx"
        file_name = "extracted/"+file_name
        df.to_excel(file_name, index=False)
        wb = load_workbook(file_name)
        ws = wb.active
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, horizontal='left', vertical='top')
        wb.save(file_name)
        print(df)
        print(file_name,"Done")