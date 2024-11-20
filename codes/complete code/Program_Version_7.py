from selenium_stealth import stealth
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import TimeoutException, NoSuchElementException,StaleElementReferenceException, WebDriverException
from selenium import webdriver
from selenium.webdriver.remote.webelement import WebElement
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from bs4 import BeautifulSoup
import re
import time
import nltk
from nltk import pos_tag, word_tokenize

def setup_driver():
    options = Options()
    options.add_argument('--disable-gpu')
    options.add_argument('--no-sandbox')
    options.add_argument('--incognito')
    options.add_argument('--lang=en-US')
    options.add_argument('--disable-blink-features=AutomationControlled')

    driver = webdriver.Chrome(options=options)
    stealth(driver,
            languages=["en-US", "en"],
            vendor="Google Inc.",
            platform="Win32",
            webgl_vendor="Intel Inc.",
            renderer="Intel Iris OpenGL Engine",
            fix_hairline=True)
    
    return driver


def get_google_search_urls(query, num_results):
    """
    Retrieves URLs from Google search results for the given query.
    Returns a list of URLs.
    """
    driver = setup_driver()
    search_urls = []
    #sites = ['autozone.com','kalpartz.com','shopadvanceautoparts.com','vanhorntruckparts.com','ebay.com','nickstruckparts.com','thewrenchmonkey.com','finditparts.com','www.amazon.com','accessorymods.com','centralalbertapaintsupply','fleetpride.com']
    sites =[]
    #required_results = (len(sites) * 2) + num_results
    try:
        # Step 1: Search each trusted site individually, taking only 2 URLs per site
        for site in sites:
            site_query = f"{query} {site}"
            driver.get("https://www.google.com")
            search_box = driver.find_element(By.NAME, "q")
            search_box.send_keys(site_query)
            search_box.send_keys(Keys.RETURN)

            # Wait a shorter time for results (5 seconds) to avoid hanging on zero-result pages
            try:
                WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, "div.yuRUbf a")))
                results = driver.find_elements(By.CSS_SELECTOR, "div.yuRUbf a")[:2]
                site_urls = [result.get_attribute("href") for result in results]
                search_urls.extend(site_urls)
            except Exception as e:
                print(f"No results for {site}, moving to the next site.")

            # Stop if we reach the required results
            if len(search_urls) >= num_results:
                break


        # Step 2: Perform a general search if not enough results from trusted sites
        if len(search_urls) < num_results:
            general_query = f"{query}"
            driver.get("https://www.google.com")
            search_box = driver.find_element(By.NAME, "q")
            search_box.send_keys(general_query)
            search_box.send_keys(Keys.RETURN)
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, "div.yuRUbf a")))

            # Extract URLs from the first page of the general search
            results = driver.find_elements(By.CSS_SELECTOR, "div.yuRUbf a")
            general_urls = [result.get_attribute("href") for result in results]
            search_urls.extend(general_urls)

            # Fetch additional pages if needed
            while len(search_urls) < num_results:
                try:
                    next_button = driver.find_element(By.ID, "pnnext")
                    next_button.click()
                    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, "div.yuRUbf a")))
                    results = driver.find_elements(By.CSS_SELECTOR, "div.yuRUbf a")
                    search_urls.extend([result.get_attribute("href") for result in results])
                except Exception as e:
                    print(f"Error navigating to the next page: Endind of the list")
                    break

    finally:
        driver.quit()
    print(num_results)
    return search_urls


def extract_data(patterns,text):
    extracted = {}
    for key, pattern in patterns.items():
        match = re.search(pattern, str(text))  # Search for the pattern in the text
        if match:
            # Check if group(1) exists to avoid IndexError
            if match.lastindex and match.group(1):
                extracted[key] = match.group(1).strip()
    return extracted


def site_specific_extraction(url, part_column, cross_ref_column):
    patterns = {
    'part_number': r'\b[A-Z]{1,4}[-:]?\d{3,6}\b(?![^/]*\/)',  # Stop at '/'
    'sku': r'\s*SKU\s*[:#]?\s*([A-Za-z0-9-]+)(?![^/]*\/)',  # Example: SKU : K440027 or SKU#K440027
    'MPN': r'\s*MPN\s*[:#]?\s*([A-Za-z0-9-]+)(?![^/]*\/)',  # Example: MPN: K440027
    'manufacturer_part_number': r'Manufacturer Part Number\s*[:#]?\s*([A-Za-z0-9-]+)(?![^/]*\/)',  
    'same_as': r'Same As\s*([A-Za-z0-9,\s.-]+?)(?![^/]*\/)',  
    'part_interchanges': r'Part Interchanges\s*([A-Za-z0-9,\s.-]+?)(?![^/]*\/)',  
    'oe_numbers': r'OE Numbers\s*([A-Za-z0-9,\s.-]+?)(?![^/]*\/)',  
    'oe_cross_reference': r'OE Cross Reference\s*([A-Za-z0-9,\s.-]+?)(?![^/]*\/)',  
}

    
    if 'fleetpride' in url:
        # Example: Extract entire part number and cross-reference from Amazon
        return part_column,cross_ref_column
    elif 'baltimoreauto' in url:
        # Example: Extract data directly without pattern-matching
        return part_column,cross_ref_column
    else:
        # Default: Use pattern-based extraction
        part_data = extract_data(patterns, part_column)
        cross_ref_data = extract_data(patterns, cross_ref_column)

        return (
        ', '.join([f'{k}: {v}' for k, v in part_data.items()]) if part_data else None,
        ', '.join([f'{k}: {v}' for k, v in cross_ref_data.items()]) if cross_ref_data else None)

def mrk(df):
    filtered_df = df.dropna(subset=['Part_No from site', 'Cross Reference'], how='all')
    return filtered_df

def extract_cross_references(text):
    # Keywords to search for cross-referencing
    keywords = re.compile(r"\b(cross|replaces|interchange|oem|part number|sku|part|partnumber)\b", re.IGNORECASE)
    
    # Regex to capture alphanumeric part numbers or SKU patterns (e.g., '4368', 'C4368', 'SKU4368')
    part_number_pattern = re.compile(r'\b[A-Z0-9]+(?:[-\s]?[A-Z0-9]+)*\b')

    cross_refs = []
    
    # Split sentences by keywords and search for part numbers following keywords
    for match in keywords.finditer(text):
        # Find position of the keyword and extract the substring starting from there
        start_index = match.end()
        substring = text[start_index:]

        # Find part numbers within the substring
        parts = part_number_pattern.findall(substring)
        
        # Append found part numbers with associated keyword context
        if parts:
            keyword = match.group().lower()
            cross_refs.append((keyword, ', '.join(parts)))

    return cross_refs

def price_comparison(query, num_results, part_num, description):
    """
    Performs price comparison for the given query by extracting product details from multiple websites.
    Returns a pandas DataFrame containing the product details.
    """
    driver = setup_driver()
    nltk.download('punkt')
    nltk.download('averaged_perceptron_tagger')
    urls = get_google_search_urls(query, num_results)
    # #urls_list = urls_list.tolist()

    # urls = list(set(urls_list))
    #print(urls)
    #urls = ['https://www.continentalbattery.com/product-reference/advance-diehard-silver-561','https://www.amazon.com/Dorman-485-701-Grease-Fitting/dp/B0036C9DR0','https://excofilter.com/description/AX230207/93800-CARQUEST']
    all_product_details = []
    #urls = [' https://nickstruckparts.com/products/coolantelbow-561-17275-polyester-reinforced-45-deg?srsltid=AfmBOoqwPg0RDxkqbjqBBDDZBwMLwX5sUIjiF4UZarahHiC0qeQUOsXk']
    #urls = ['https://partsavatar.ca/moog-pitman-arm-k440027?srsltid=AfmBOoqlA3wn4msup9cKc1HZPcxEBtLqPOpCIvBb6YBYZ-qPP-E1rHmT']
    for url in urls:
        try:
            #print(class_names)
            # Gather product details
            details = {
            'Part Num to search': part_num,
            'Description' : description,
            'url': url,
            'Part_No from site': extract_product_details(url,['part detail title number','product-line-sku','product_info__description_list','part_number', 'partNumSection', 'part-number', 'product_part-info', 'item-part-number','sku','sku number','item part number','part number','item no','product-basic-details__text--part'],driver),
            'Other Details': extract_product_details(url,['availability', 'stock', 'in-stock','warranty', 'guarantee','spec', 'specs', 'details', 'product-spec','specification','tech','item-description','description','prodDetails', 'ProductDetail', 'item-desc', 'product_description', 'techSpec', 'item-description', 'product-details', 'product_info', 'description','corePrice', 'product-price', 'price-primary','item price','productName', 'product-title', 'mainTitle', 'product-detail-title','product-basic-details__text--title', 'productTitle', 'titleSection','Product header','product title','item tilte','item name','product name','product-details-information','OECROSSREFERENCE','product-specifications','interchange category-description','cross-reference','other-cross-refrence','product-description rte','productDescription','crossRefWrapper','interchangeItemValue','productDetails_techSpec_section_1','product-details-d','css-cm8roc','product_description__row','product-part-interchange','product-specifications-container','cross-reference-list','product-single__description','x-item-description-child','interchange','Crossreference','replaces', 'product superseded','superseded','interchages'],driver)
            }
            #print(details)
            all_product_details.append(details)
            print(f"Website done: {url}")

        except Exception as e:
            print(f"Error processing URL {url}: {e}")
    driver.quit()

    for detail in all_product_details:
        cross_refs = extract_cross_references(detail['Other Details'])
        detail['Extracted Cross References'] = cross_refs

    
    df= pd.DataFrame(all_product_details)
    #df = mrk(df_unfiltered)
    
    #df[['Extracted_Part No From Site', 'Extracted_Cross Reference']] = df.apply(lambda row: site_specific_extraction(row['url'], row['Part_No from site'], row['Other Details']),axis=1, result_type='expand')
    return df

def normalize_string(s):
        """Helper function to normalize strings by removing spaces, dashes, and underscores."""
        return re.sub(r'[\s+\-\_+]', '', s).lower()

def extract_data_from_elements(elements, substrings_set, url, driver):
    """Helper function to extract data from collapsible elements."""
    extracted_data = []
    for element in elements:
        element_classes = [cls.lower() for cls in element.get('class', [])]
        element_id = element.get('id', '').lower()

        # Check if the element matches based on class or ID
        if any(substr in normalize_string(cls) for cls in element_classes + [element_id] for substr in substrings_set):
            extracted_text = ' '.join(element.stripped_strings)
            if extracted_text:
                extracted_data.append(extracted_text)
            if 'partsavatar' in url and element_id == 'oecrossreference':
                try:
                    clickable_div = WebDriverWait(driver, 10).until(
                        EC.presence_of_element_located((By.ID, "OECROSSREFERENCE"))
                    )

                    # Scroll element into view properly
                    driver.execute_script(
                        "arguments[0].scrollIntoView({block: 'center', inline: 'nearest'});", 
                        clickable_div
                    )
                    time.sleep(1)  # Optional: Brief pause for smooth scrolling

                    # Wait for clickability and perform click
                    WebDriverWait(driver, 10).until(
                        EC.element_to_be_clickable((By.ID, "OECROSSREFERENCE"))
                    ).click()

                    time.sleep(2)
                    extracted_txt = clickable_div.text.strip()
                    if extracted_txt:
                        extracted_data.append(extracted_txt)
                        print("Extracted Text from Clickable Div:")
                        #print(extracted_txt)

                except Exception as e:
                    print(f"Error interacting with the element: {e}")
    return ' /'.join(extracted_data)


def extract_product_details(url, substrings, driver):
        # Navigate to the URL
    driver.get(url)
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.TAG_NAME, 'body')))

    # Normalize substrings in advance to avoid repeated computation
    substrings_set = {normalize_string(substr) for substr in substrings}

    # Extract data from the main page
    main_soup = BeautifulSoup(driver.page_source, 'html.parser')
    extracted_data = extract_data_from_elements(main_soup.find_all(), substrings_set, url, driver)

    for iframe in driver.find_elements(By.TAG_NAME, 'iframe'):
        try:
            driver.switch_to.frame(iframe)
            iframe_soup = BeautifulSoup(driver.page_source, 'html.parser')
            iframe_data = extract_data_from_elements(iframe_soup.find_all(), substrings_set, url, driver)
            
            # If iframe_data is not empty, extend the main extracted data
            if iframe_data:
                #extracted_data.join(' ' + iframe_data)
                extracted_data += ' ' + iframe_data
        except Exception as e:
            print(f"Error processing iframe: {e}")
        finally:
            driver.switch_to.default_content()

    return extracted_data

if __name__ == "__main__":
    location = r"C:\Users\MRK\Downloads\20 Parts run wo prior web & line.xlsx"
    list = pd.read_excel(location)
    #print(list)
    combined_list = list.apply(lambda row: f"{row['Part Num to search']} {row['Description']}", axis=1).tolist()
    for i in range(20):
        desc = list.iloc[i]['Description'] 
        partnum = list.iloc[i]['Part Num to search']
        query = combined_list[i]
        print(query)
        df = price_comparison(query, num_results=20, part_num = partnum , description = desc)
        file_name = f"{query.replace(' ', '_')}_project_part_details.xlsx"
        df.to_excel(file_name, index=False)
        wb = load_workbook(file_name)
        ws = wb.active
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, horizontal='left', vertical='top')
        wb.save(file_name)
        print(df)
        print(file_name,"Done")