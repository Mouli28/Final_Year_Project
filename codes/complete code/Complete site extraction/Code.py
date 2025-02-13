from selenium_stealth import stealth
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import TimeoutException, NoSuchElementException,StaleElementReferenceException, WebDriverException
from selenium import webdriver
import undetected_chromedriver as uc
from selenium.webdriver.remote.webelement import WebElement
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from bs4 import BeautifulSoup
import re
import random
import time
import nltk
from nltk import pos_tag, word_tokenize


def human_delay(min_delay=1, max_delay=5):
    time.sleep(random.uniform(min_delay, max_delay))

def simulate_user_interaction(driver, element):
    actions = ActionChains(driver)
    actions.move_to_element(element).perform()
    human_delay()
    actions.click().perform()


def setup_driver():
    """ Setup Undetected Selenium WebDriver """
    user_agents = [
        'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/110.0.0.0 Safari/537.36',
        'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/110.0.0.0 Safari/537.36',
        'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/109.0.0.0 Safari/537.36'
    ]
    user_agent = random.choice(user_agents)

    options = uc.ChromeOptions()
    options.add_argument(f"user-agent={user_agent}")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-gpu")
    options.add_argument("--incognito")
    options.add_argument("--disable-extensions")

    driver = uc.Chrome(options=options, headless=False)  # Keep headless=False to look human

    return driver

def get_google_search_urls(query, num_results):
    """
    Retrieves URLs from Google search results for the given query.
    Returns a list of URLs.
    
    """
    driver = setup_driver()
    search_urls = []
    #sites = ['autozone.com','kalpartz.com','shopadvanceautoparts.com','vanhorntruckparts.com','ebay.com','nickstruckparts.com','thewrenchmonkey.com','finditparts.com','www.amazon.com','accessorymods.com','centralalbertapaintsupply','fleetpride.com']
    sites =[]
    #required_results = (len(sites) * 2) + num_results
    try:
        # Step 1: Search each trusted site individually, taking only 2 URLs per site
        for site in sites:
            site_query = f"{query} {site}"
            driver.get("https://www.google.com")
            search_box = driver.find_element(By.NAME, "q")
            search_box.send_keys(site_query)
            search_box.send_keys(Keys.RETURN)

            # Wait a shorter time for results (5 seconds) to avoid hanging on zero-result pages
            try:
                WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, "div.yuRUbf a")))
                results = driver.find_elements(By.CSS_SELECTOR, "div.yuRUbf a")[:2]
                site_urls = [result.get_attribute("href") for result in results]
                search_urls.extend(site_urls)
            except Exception as e:
                print(f"No results for {site}, moving to the next site.")

            # Stop if we reach the required results
            if len(search_urls) >= num_results:
                break


        # Step 2: Perform a general search if not enough results from trusted sites
        if len(search_urls) < num_results:
            general_query = f"{query}"
            driver.get("https://www.google.com")
            search_box = driver.find_element(By.NAME, "q")
            search_box.send_keys(general_query)
            search_box.send_keys(Keys.RETURN)
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, "div.yuRUbf a")))

            # Extract URLs from the first page of the general search
            results = driver.find_elements(By.CSS_SELECTOR, "div.yuRUbf a")
            general_urls = [result.get_attribute("href") for result in results]
            search_urls.extend(general_urls)

            # Fetch additional pages if needed
            while len(search_urls) < num_results:
                try:
                    next_button = driver.find_element(By.ID, "pnnext")
                    next_button.click()
                    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, "div.yuRUbf a")))
                    results = driver.find_elements(By.CSS_SELECTOR, "div.yuRUbf a")
                    search_urls.extend([result.get_attribute("href") for result in results])
                except Exception as e:
                    print(f"Error navigating to the next page: Endind of the list")
                    break

    finally:
        driver.quit()
    print(num_results)
    return search_urls


def extract_text_from_urls(urls):
    """
    Extracts all visible text from the given list of URLs.
    Returns a dictionary where keys are URLs and values are the extracted text.
    """
    driver = setup_driver()  # Reuse your existing setup_driver function
    data = []

    for url in urls:
        random_delay = random.uniform(1, 5)
        time.sleep(random_delay)
        try:
            print(url)
            driver.get(url)
            time.sleep(3)  # Allow time for the page to load
            
            # Get the page source
            page_source = driver.page_source
            
            # Parse with BeautifulSoup
            soup = BeautifulSoup(page_source, "html.parser")
            
            # Extract visible text
            for script in soup(["script", "style"]):  # Remove script and style elements
                script.decompose()
            text = soup.get_text(separator=" ")
            
            # Clean the text
            clean_text = re.sub(r'\s+', ' ', text).strip()
            
            # Add to data list
            data.append({"URL": url, "Extracted_Text": clean_text})

        except Exception as e:
            print(f"Error extracting text from {url}: {e}")
            data.append({"URL": url, "Extracted_Text": None})  # Add None if extraction fails

    driver.quit()
    
    # Convert to DataFrame
    df = pd.DataFrame(data)
    return df


if __name__ == "__main__":
    location = r"C:\Users\MRK\Downloads\20 Parts run wo prior web & line.xlsx"
    list = pd.read_excel(location)
    #print(list)
    combined_list = list.apply(lambda row: f"{row['Part Num to search']} {row['Description']}", axis=1).tolist()
    for i in range(1):
        desc = list.iloc[i]['Description'] 
        partnum = list.iloc[i]['Part Num to search']
        query = combined_list[i]
        print(query)
        urls = get_google_search_urls(query, 20)
        df = extract_text_from_urls(urls)
        file_name = f"{query.replace(' ', '_')}_project_part_details.xlsx"
        df.to_excel(file_name, index=False)
        wb = load_workbook(file_name)
        ws = wb.active
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, horizontal='left', vertical='top')
        wb.save(file_name)
        print(df)
        print(file_name,"Done")
