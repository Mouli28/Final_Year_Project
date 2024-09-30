from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
import pandas as pd
from transformers import pipeline
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from bs4 import BeautifulSoup
import re
import time

# Function to extract Google search URLs
def get_google_search_urls(query, num_results):
    """
    Retrieves URLs from Google search results for the given query.

    Parameters:
    - query (str): The search query.
    - num_results (int): Number of search results to retrieve.

    Returns:
    - list: A list of URLs extracted from the search results.
    """
    search_urls = []
    options = Options()
    options.add_argument('--disable-gpu')
    options.add_argument('--no-sandbox')
    options.add_argument('--incognito')
    driver = webdriver.Chrome(options=options)
    driver.get("https://www.google.com")

    # Locate the search box, enter the query, and perform the search
    search_box = driver.find_element(By.NAME, "q")
    search_box.send_keys(query)
    search_box.send_keys(Keys.RETURN)
    
    # Wait for results to load
    driver.implicitly_wait(2)

    urls = []

    # Extract URLs from the first page
    results = driver.find_elements(By.CSS_SELECTOR, "div.yuRUbf a")
    urls.extend([result.get_attribute("href") for result in results])

    # If less than the required number of results, navigate to the next page
    while len(urls) < num_results:
        try:
            next_button = driver.find_element(By.ID, "pnnext")
            next_button.click()
            driver.implicitly_wait(2)
            results = driver.find_elements(By.CSS_SELECTOR, "div.yuRUbf a")
            urls.extend([result.get_attribute("href") for result in results])
        except Exception as e:
            print(f"Error navigating to the next page: {e}")
            break

    driver.quit()

    # Return only the required number of URLs
    return urls[:num_results]


# Function to extract product details from a specific webpage
def extract_product_details(url, substrings):
    """
    Extracts product details from the given URL by dynamically identifying
    class names containing specific substrings and extracting all instances
    of data for each matched class.

    Parameters:
    - url (str): The URL of the product page.
    - substrings (dict): A dictionary where keys are the data fields (e.g., 'name', 'price')
                         and values are lists of substrings to search for in class names.

    Returns:
    - dict: A dictionary containing the extracted product details.
    """
    # Initialize the dictionary to store product details
    product_details = {}

    options = Options()
    options.add_argument('--disable-gpu')
    options.add_argument('--no-sandbox')
    options.add_argument('--incognito')
    driver = webdriver.Chrome(options=options)

    try:
        driver.get(url)
        # Wait until the page is fully loaded
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.TAG_NAME, 'body')))
        page_source = driver.page_source
    except Exception as e:
        print(f"Error fetching the URL {url}: {e}")
        driver.quit()
        return product_details

    driver.quit()

    # Parse the page source with BeautifulSoup
    soup = BeautifulSoup(page_source, 'html.parser')

    # Extract all class names
    all_elements = soup.find_all(True, class_=True)
    class_names = set()
    for element in all_elements:
        classes = element.get('class')
        for cls in classes:
            class_names.add(cls)

    # For each data field, find all class names that contain any of the specified substrings
    for field, substr_list in substrings.items():
        matched_classes = []
        for cls in class_names:
            for substr in substr_list:
                if substr.lower() in cls.lower():
                    matched_classes.append(cls)
                    break

        # If we found matched classes, extract data from all instances
        if matched_classes:
            extracted_data = []
            for matched_class in matched_classes:
                elements = soup.find_all(class_=re.compile(matched_class))
                for element in elements:
                    extracted_data.append(element.get_text(strip=True))
            # Store all instances in a list, joining them as a single string for display
            product_details[field] = '; '.join(extracted_data)
        else:
            product_details[field] = None

    return product_details

# Function to perform price comparison using extracted product details
def price_comparison(query, num_results):
    """
    Performs price comparison for the given query by extracting product details from multiple websites.

    Parameters:
    - query (str): The search query (e.g., "AD1066 bracket").
    - num_results (int): Number of Google search results to process.

    Returns:
    - tuple: A pandas DataFrame containing the product details and a list of URLs processed.
    """
    urls = get_google_search_urls(query, num_results)
    all_product_details = []
    processed_urls = []

    # Define substrings for dynamic class name matching
    substrings = {
        'Product name': ['productName', 'product-title', 'mainTitle', 'product-detail-title', 'productTitle', 'titleSection'],
        'price': ['corePriceDisplay_desktop_feature_div', 'price', 'msrp', 'product-price', 'x-price-primary'],
        'description': ['vim d-item-description','description','prodDetails', 'ProductDetails', 'item-desc isColorImage', 'product_description_wrapper', 'productDetails_techSpec_section_1', 'x-item-description-child', 'product-details', 'product_info_description_list','d-item-description','product-details-inner', 'description-collapse', 'description','product-details-module'],
        'Taxonomy' : ['breadcrumb', 'bread-crumb'],
        'Part No': ['part_number', 'partNumSection', 'part-number', 'product_part-info', 'item-part-number'],
        'Cross Reference':['Crossreference','replaces', 'Interchange', 'product superseded','superseded','x-item-description-child'],
        'specifications': ['spec', 'specs', 'details', 'product-spec','specification-collapse'],
        'warranty': ['warranty', 'guarantee'],
        'availability': ['availability', 'stock', 'in-stock']
    }

    for url in urls:
        details = extract_product_details(url, substrings)
        if details:
            details['url'] = url  # Optionally add the URL to the details
            all_product_details.append(details)
            processed_urls.append(url)

    df = pd.DataFrame(all_product_details)
    return df, processed_urls


if __name__ == "__main__":
    location = filelocation here #file location here
    list = pd.read_excel(location)
    combined_list = list.apply(lambda row: f"{row['Description']} {row['Part Number']}", axis=1).tolist()
    for query in combined_list:
        df, urls = price_comparison(query, num_results=50)
        file_name = f"{query.replace(' ', '_')}_nlp_part_details.xlsx"
        df.to_excel(file_name, index=False)
        wb = load_workbook(file_name)
        ws = wb.active
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, horizontal='left', vertical='top')
        wb.save(file_name)
    print(df)

