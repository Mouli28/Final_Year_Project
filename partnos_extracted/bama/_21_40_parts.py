from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from bs4 import BeautifulSoup
import requests
from googlesearch import search
import time
import random
import re

def get_google_search_urls(query, num_results=20):
    driver = init_driver()
    driver.get("https://www.google.com")

    # Locate the search box, enter the query, and perform the search
    search_box = driver.find_element(By.NAME, "q")
    search_box.send_keys(query)
    search_box.send_keys(Keys.RETURN)
    
    # Wait for results to load
    driver.implicitly_wait(2)

    urls = []

    # Extract URLs from the first page
    results = driver.find_elements(By.CSS_SELECTOR, "div.yuRUbf a")
    urls.extend([result.get_attribute("href") for result in results])

    # If less than the required number of results, navigate to the next page
    while len(urls) < num_results:
        try:
            next_button = driver.find_element(By.ID, "pnnext")
            next_button.click()
            driver.implicitly_wait(2)
            results = driver.find_elements(By.CSS_SELECTOR, "div.yuRUbf a")
            urls.extend([result.get_attribute("href") for result in results])
        except Exception as e:
            print(f"Error navigating to the next page: {e}")
            break

    driver.quit()

    # Return only the required number of URLs
    return urls[:num_results]


def init_driver():
    options = Options()
    options.add_argument('--disable-gpu')
    options.add_argument('--no-sandbox')
    options.add_argument('--incognito')
    driver = webdriver.Chrome(options=options)
    return driver

def price_comparison(query):
    driver = init_driver()
    urls = get_google_search_urls(query)
    product_details = []

    for url in urls:
        details = extract_product_details(driver, url)
        if details:
            product_details.append(details)

    driver.quit()
    df = pd.DataFrame(product_details)
    return df, urls

def extract_with_class_names(soup, tag_list, class_list):
    for tag in tag_list:
        elements = soup.find_all(tag)
        for element in elements:
            if any(cls in element.get('id', '') for cls in class_list):
                return re.sub(r'\s+', ' ', element.get_text()).strip()
            if any(cls in ' '.join(element.get('class', [])) for cls in class_list):
                return re.sub(r'\s+', ' ', element.get_text()).strip()
            else:
                    for cls in class_list:
                        if element.get('data-pl') == cls:
                            return re.sub(r'\s+', ' ', element.get_text()).strip()
    return None



def extract_product_details(driver, url):
    try:
        driver.get(url)
        WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.TAG_NAME, 'body')))
        time.sleep(3)  # Additional sleep for dynamic content to load
        
        soup = BeautifulSoup(driver.page_source, 'html.parser')

        details = {
            'url': url,
            'Product Name': extract_with_class_names(soup, ['h1', 'p', 'h3', 'div'], 
                                                     ["stellar_phase_2","productName","product-title","x-item-title__mainTitle","product-detail-title","productTitle","titleSection","product_title entry-title"]),
            'Price': extract_with_class_names(soup, ['p', 'div', 'span'], 
                                               ["corePriceDisplay_desktop_feature_div","price", "msrp", "product-price","x-price-primary"]),
            'Part No': extract_with_class_names(soup, ['li', 'div', 'span'], 
                                                ["part_number","partNumSection","x-item-description-child","part-number","product_part-info","item-part-number"]),
            'Taxonomy': extract_with_class_names(soup, ['div', 'a', 'ol'], 
                                                 ["showing-breadcrumbs_div","page-bread-crumbs","breadcrumb","bread-crumbs","breadcrumnb","seo-breadcrumb-text","breadcrumbs","breadcrumb-nav","site-breadcrumb js-site-breadcrumb","breadcrumb-container"]),
            'Cross Reference': extract_with_class_names(soup, ['ul', 'span'], 
                                                        ["list-unstyled cross-reference-list","body-3 alt-stock-code-text","Crossreference","replaces","Interchange","product-superseded-list"]),
            'Details': extract_with_class_names(soup, ['p', 'div', 'table'], 
                                                ["description","prodDetails","ProductDetails","whyBuyThis","item-desc isColorImage","product_description_wrapper","productDetails_techSpec_section_1","x-item-description-child","product-details","product_info_description_list","product-details-inner","tab-6","description-collapse","product-details-module"]),
            'Specification': extract_with_class_names(soup, ['p', 'div', 'table'], 
                                                ["specification-collapse","productDetails_db_sections","vim x-about-this-item"]),
            'Warranty': extract_with_class_names(soup, ['p', 'div'], 
                                                 ['WarrantyInfo-collapse', 'warranty']),
            'Availability': extract_with_class_names(soup, ['p', 'div'], ["availability","productAvailability-Outofstock","outofstock"])
        }
        
        print(details)
        return details if any(details.values()) else None

    except Exception as e:
        print(f"Error extracting data from {url}: {e}")
        return None

if __name__ == "__main__":
    parts = ['2357 Light Bulb', '4114 Light Bulb', '7440 Light Bulb', '7441 Mini Bulb', '1156NA Light Bulb', '194BP License Plate Light Bulb', '211-2 Light Bulb', '2357A Light Bulb', '3057NA Light Bulb', '3156NA Light Bulb', '3457NA Light Bulb', '7440NA Light Bulb', '9003BP Halogen Bulb', '9004BP Headlamp Bulb', '9005BP Halogen Bulb', '9006BP Halogen Bulb', '9007BP Halogen Bulb', '9008BP Halogen Bulb', '916NA Light Bulb', 'C24 Light Bulb']
    direc_name = "extracted/"
    for part in parts:
        query = part
        df, urls = price_comparison(query) 
        file_name = f"{query.replace(' ', '_')}_part_details.xlsx"
        df.to_excel(file_name, index=False)
        wb = load_workbook(file_name)
        ws = wb.active
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, horizontal='left', vertical = 'top')
        save_location = direc_name+file_name
        wb.save(save_location) 
    print(df)
